import os
import re
import csv
import shutil
import sqlite3
import unicodedata
from decimal import Decimal
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from ttkbootstrap.toast import ToastNotification
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as PDFImage
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.graphics.shapes import Drawing, Line
from PIL import Image, ImageTk
from reportlab.platypus import Table, TableStyle, Image as PDFImage, Paragraph, Spacer
from reportlab.lib import colors
import requests

def formatar_cnpj(cnpj):
    c = ''.join(filter(str.isdigit, str(cnpj)))
    if len(c) == 14:
        return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"
    return c
def formatar_cep(cep):
    c = ''.join(filter(str.isdigit, str(cep)))
    if len(c) == 8:
        return f"{c[:5]}-{c[5:]}"
    return c
def formatar_telefone(tel):
    t = ''.join(filter(str.isdigit, str(tel)))
    if len(t) == 11:
        return f"({t[:2]}) {t[2:7]}-{t[7:]}"
    elif len(t) == 10:
        return f"({t[:2]}) {t[2:6]}-{t[6:]}"
    return tel
def formatar_moeda(valor):
    try:
        if valor is None:
            v = 0.0
        elif isinstance(valor, str):
            s = valor.strip().replace('R$', '').replace(' ', '')
            if s == '':
                v = 0.0
            else:
                # Normaliza milhares e decimais
                if ',' in s and '.' in s:
                    s = s.replace('.', '').replace(',', '.')
                else:
                    s = s.replace('.', '').replace(',', '.')
                v = float(s)
        else:
            v = float(valor)
    except Exception:
        v = 0.0
    texto = f"{v:,.2f}"            # ex: "1234.56" -> "1,234.56"
    texto = texto.replace(",", "V").replace(".", ",").replace("V", ".")  # -> "1.234,56"
    return f"R$ {texto}"

def normalizar_chave(texto):
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto)
                    if unicodedata.category(c) != 'Mn')
    texto = texto.lower().replace(':', '').replace('%', '').replace('(', '').replace(')', '').strip()
    texto = texto.replace(' ', '_')
    return texto

class SistemaPedidos:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Orçamentos")
        self.root.geometry("1200x700")
        
        # Inicializar banco de dados
        self.init_db()
        
        # Criar notebook (abas)
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # criar estruturas temporárias
        self.itens_pedido_temp = []
        
        # Criar abas
        self.criar_aba_clientes()
        self.criar_aba_produtos()
        self.criar_aba_pedidos()
        self.criar_aba_consulta_orcamentos()
        self.criar_aba_empresas()
        self.edicao_numero_pedido = None

         # Atalhos de teclado
        self.root.bind("<Control-n>", lambda e: self.limpar_pedido())      # Novo Orçamento
        self.root.bind("<Control-s>", lambda e: self.finalizar_pedido())   # Salvar Orçamento
         
    def init_db(self):
        self.conn = sqlite3.connect('pedidos.db')
        self.cursor = self.conn.cursor()
        # clientes
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS clientes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                razao_social TEXT NOT NULL,
                cnpj TEXT UNIQUE NOT NULL,
                ie TEXT,
                endereco TEXT,
                cidade TEXT,
                estado TEXT,
                cep TEXT,
                telefone TEXT,
                email TEXT
            )
        ''')
        try:
            self.cursor.execute("ALTER TABLE pedidos ADD COLUMN empresa_id INTEGER")
            self.conn.commit()
        except Exception:
            # se já existir, ignora o erro
            pass
        # produtos
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS produtos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT UNIQUE NOT NULL,
                descricao TEXT NOT NULL,
                valor_unitario REAL NOT NULL,
                tipo TEXT,
                origem_tributacao TEXT,
                voltagem TEXT,
                aliq_icms REAL DEFAULT 0,
                aliq_ipi REAL DEFAULT 0,
                aliq_pis REAL DEFAULT 0,
                aliq_cofins REAL DEFAULT 0
            )
        ''')
        # empresas (empresas que emitem orçamentos)
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS empresas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL,
                cnpj TEXT UNIQUE,
                ie TEXT,
                endereco TEXT,
                cidade TEXT,
                estado TEXT,
                cep TEXT,
                telefone TEXT,
                email TEXT,
                caminho_logo TEXT,
                eh_padrao INTEGER DEFAULT 0
            )
        ''')

        # pedidos
        # pedidos
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS pedidos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero_pedido TEXT UNIQUE NOT NULL,
                data_pedido TEXT NOT NULL,
                cliente_id INTEGER NOT NULL,
                valor_produtos REAL,
                valor_icms REAL,
                valor_ipi REAL,
                valor_pis REAL,
                valor_cofins REAL,
                valor_total REAL,
                representante TEXT,
                condicoes_pagamento TEXT,
                desconto REAL DEFAULT 0,
                status TEXT DEFAULT 'Em Aberto',
                observacoes TEXT,
                validade TEXT,
                FOREIGN KEY (cliente_id) REFERENCES clientes (id)
            )
        ''')

        # itens_pedido
        # Itens do pedido (ligados ao número do orçamento)
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS pedido_itens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero_pedido TEXT NOT NULL,
                produto_id INTEGER NOT NULL,
                qtd REAL NOT NULL,
                valor_unitario REAL NOT NULL,
                FOREIGN KEY (numero_pedido) REFERENCES pedidos (numero_pedido),
                FOREIGN KEY (produto_id) REFERENCES produtos (id)
            )
        ''')

        self.conn.commit()
    
    def calcular_totais(self, itens):
        subtotal = sum(i['qtd'] * i['valor'] for i in itens)
        total_icms = total_ipi = total_pis = total_cofins = 0
        for item in itens:
            self.cursor.execute('SELECT aliq_icms, aliq_ipi, aliq_pis, aliq_cofins FROM produtos WHERE id=?', (item['produto_id'],))
            icms, ipi, pis, cofins = self.cursor.fetchone()
            base = item['qtd'] * item['valor']
            total_icms += base * (icms or 0) / 100
            total_ipi += base * (ipi or 0) / 100
            total_pis += base * (pis or 0) / 100
            total_cofins += base * (cofins or 0) / 100
        total = subtotal + total_icms + total_ipi + total_pis + total_cofins
        return subtotal, total_icms, total_ipi, total_pis, total_cofins, total
    def copiar_celula_treeview(self, event):
        """Copia o conteúdo da célula selecionada da Treeview (Ctrl+C ou Ctrl+Duplo Clique)."""
        tree = event.widget
        sel = tree.selection()
        if not sel:
            return

        item_id = sel[0]
        valores = tree.item(item_id, "values")
        if not valores:
            return

        # tenta pegar coluna atual
        x, y = event.x, event.y
        coluna_id = tree.identify_column(x)
        try:
            col_index = int(coluna_id.replace('#', '')) - 1
        except:
            col_index = 0

        texto = str(valores[col_index])
        self.root.clipboard_clear()
        self.root.clipboard_append(texto)
        self.root.update()

        # feedback visual leve (toast)
        try:
            
            ToastNotification(
                title="Copiado!",
                message=f"'{texto}' copiado para a área de transferência.",
                duration=1500,
                bootstyle="info"
            ).show_toast()
        except Exception:
            pass

    def criar_aba_empresas(self):
        aba = ttk.Frame(self.notebook)
        self.notebook.add(aba, text="Empresas")
        # Botões de ação
        frame_botoes = ttk.Frame(aba)
        frame_botoes.pack(fill="x", padx=10, pady=5)
        ttk.Button(frame_botoes, text="Adicionar", command=self.adicionar_empresa).pack(side="left", padx=5)
        ttk.Button(frame_botoes, text="Editar", command=self.editar_empresa).pack(side="left", padx=5)
        ttk.Button(frame_botoes, text="Excluir", command=self.excluir_empresa).pack(side="left", padx=5)

        # Lista de empresas
        frame_lista = ttk.Frame(aba)
        frame_lista.pack(fill="both", expand=True, padx=10, pady=10)

        colunas = ("ID", "Nome", "CNPJ", "Cidade", "Telefone")
        self.tree_empresas = ttk.Treeview(frame_lista, columns=colunas, show="headings")
        for col in colunas:
            self.tree_empresas.heading(col, text=col)
            self.tree_empresas.column(col, width=150, anchor="center")
        self.tree_empresas.pack(fill="both", expand=True)

        self.carregar_empresas()        

    def buscar_cep(self, cep, entries):
        """Busca endereço via API e preenche campos."""
        cep = cep.strip().replace("-", "")
        if len(cep) != 8 or not cep.isdigit():
            messagebox.showwarning("CEP inválido", "Digite um CEP válido com 8 dígitos.")
            return

        try:
            resposta = requests.get(f"https://viacep.com.br/ws/{cep}/json/")
            if resposta.status_code == 200:
                dados = resposta.json()
                if "erro" in dados:
                    messagebox.showwarning("Aviso", "CEP não encontrado.")
                    return
                # Preencher campos automaticamente
                entries["endereco"].delete(0, tk.END)
                entries["endereco"].insert(0, dados.get("logradouro", ""))
                entries["cidade"].delete(0, tk.END)
                entries["cidade"].insert(0, dados.get("localidade", ""))
                entries["estado"].delete(0, tk.END)
                entries["estado"].insert(0, dados.get("uf", ""))
            else:
                messagebox.showerror("Erro", "Não foi possível buscar o CEP.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao buscar CEP: {e}")

    def criar_aba_clientes(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Clientes")

        # --- Barra de pesquisa ---
        search_frame = ttk.Frame(frame)
        search_frame.pack(fill='x', padx=10, pady=5)

        ttk.Label(search_frame, text="Pesquisar:").pack(side='left', padx=5)
        self.entry_pesquisa_cliente = ttk.Entry(search_frame, width=40)
        self.entry_pesquisa_cliente.pack(side='left', padx=5)
        
        ttk.Button(search_frame, text="Buscar", command=lambda: self.carregar_clientes(self.entry_pesquisa_cliente.get())).pack(side='left', padx=5)
        ttk.Button(search_frame, text="Limpar", command=lambda: self.carregar_clientes()).pack(side='left', padx=5)

        # --- Barra de botões ---
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(btn_frame, text="Adicionar", bootstyle=SUCCESS, command=self.adicionar_cliente).pack(side='left', padx=5 )
        ttk.Button(btn_frame, text="Editar", bootstyle=INFO, command=lambda: self.editar_cliente(None)).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Excluir", bootstyle=DANGER, command=self.excluir_cliente).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Importar Arquivo", bootstyle=WARNING, command=lambda: self.importar_dados("clientes")).pack(side='left', padx=5)

        # --- Lista de clientes ---nu
        list_frame = ttk.LabelFrame(frame, text="Clientes Cadastrados", padding=10)
        list_frame.pack(fill='both', expand=True, padx=10, pady=10)

        cols = ('Razão Social', 'CNPJ', 'Cidade', 'Telefone')
        self.tree_clientes = ttk.Treeview(list_frame, columns=cols, show='headings', height=12)
        for col in cols:
            self.tree_clientes.heading(col, text=col)
            self.tree_clientes.column(col, width=140, anchor='center')

        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.tree_clientes.yview)
        self.tree_clientes.configure(yscrollcommand=scrollbar.set)
        self.tree_clientes.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

                # Permitir copiar célula
        self.tree_clientes.bind("<Control-c>", self.copiar_celula_treeview)
        self.tree_clientes.bind("<Control-Double-1>", self.copiar_celula_treeview, add='+')


        # Duplo clique abre visualização
        self.tree_clientes.bind("<Double-1>", self.visualizar_cliente)

        self.carregar_clientes()

    def abrir_formulario_cliente(self, cliente=None):
        """Abre popup para adicionar/editar cliente."""
        top = tk.Toplevel(self.root)
        top.title("Cadastro de Cliente")
        top.geometry("600x400")

        # deixa responsivo
        top.grid_columnconfigure(1, weight=1)
        top.grid_columnconfigure(3, weight=1)

        labels = ['Nome:', 'CNPJ:', 'IE:', 'CEP:', 'Endereço:', 'Cidade:', 'Estado:', 'Telefone:', 'Email:']
        entries = {}
        for i, label in enumerate(labels):
            ttk.Label(top, text=label).grid(row=i//2, column=(i%2)*2, sticky='w', padx=5, pady=5)
            entry = ttk.Entry(top, width=28)
            entry.grid(row=i//2, column=(i%2)*2+1, padx=5, pady=5, sticky="ew")
            chave = normalizar_chave(label)
            entries[chave] = entry

        # Se for edição, preencher os campos
        if cliente:
            keys = ['id','razao_social','cnpj','ie', 'cep', 'endereco','cidade','estado','telefone','email']
            for k, v in zip(keys, cliente):
                if k in entries:
                    entries[k].insert(0, v or "")
        
        def salvar():
            dados = {k: v.get().strip() for k, v in entries.items()}
            campos_obrigatorios = {"razao_social": "Razão Social", "cnpj": "CNPJ"}
            faltando = [nome for chave, nome in campos_obrigatorios.items() if not dados.get(chave)]

            if faltando:
                messagebox.showwarning("Atenção", f"Preencha os campos obrigatórios: {', '.join(faltando)}")
                return

            try:
                if cliente:  # editar
                    self.cursor.execute('''
                        UPDATE clientes
                        SET razao_social=?, cnpj=?,cep=?, ie=?, endereco=?, cidade=?, estado=?, cep=?, telefone=?, email=?
                        WHERE id=?
                    ''', (dados['razao_social'], dados['cnpj'], dados.get('ie'),dados.get('cep'), dados.get('endereco'),
                        dados.get('cidade'), dados.get('estado'),
                        dados.get('telefone'), dados.get('email'), cliente[0]))
                else:  # novo
                    self.cursor.execute('''
                        INSERT INTO clientes (razao_social, cnpj, ie, cep, endereco, cidade, estado, telefone, email)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (dados['razao_social'], dados['cnpj'], dados.get('ie'),dados.get('cep'), dados.get('endereco'),
                        dados.get('cidade'), dados.get('estado'),
                        dados.get('telefone'), dados.get('email')))
                self.conn.commit()
                self.carregar_clientes()
                top.destroy()
            except sqlite3.IntegrityError:
                messagebox.showerror("Erro", "CNPJ já cadastrado!")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar cliente: {e}")

        ttk.Button(top, text="Salvar", command=salvar).grid(row=6, column=0, columnspan=2, pady=10)
    def criar_aba_empresas(self):
            frame = ttk.Frame(self.notebook)
            self.notebook.add(frame, text="Empresas")

            # botões
            btn_frame = ttk.Frame(frame)
            btn_frame.pack(fill='x', padx=10, pady=5)
            ttk.Button(btn_frame, text="Adicionar", bootstyle=SUCCESS, command=self.abrir_formulario_empresa).pack(side='left', padx=5)
            ttk.Button(btn_frame, text="Editar", bootstyle=INFO, command=self.editar_empresa).pack(side='left', padx=5)
            ttk.Button(btn_frame, text="Excluir", bootstyle=DANGER, command=self.excluir_empresa).pack(side='left', padx=5)

            # lista
            list_frame = ttk.LabelFrame(frame, text="Empresas Cadastradas", padding=10)
            list_frame.pack(fill='both', expand=True, padx=10, pady=10)
            cols = ('Nome','CNPJ','Cidade','Telefone')
            self.tree_empresas = ttk.Treeview(list_frame, columns=cols, show='headings', height=12)
            for col in cols:
                self.tree_empresas.heading(col, text=col)
                self.tree_empresas.column(col, width=160, anchor='center')
            self.tree_empresas.pack(fill='both', expand=True)
            self.carregar_empresas()

    def adicionar_cliente(self):
        self.abrir_formulario_cliente()

    def editar_cliente(self, event=None):
        item = self.tree_clientes.selection()
        if not item:
            messagebox.showwarning("Atenção", "Selecione um cliente para editar.")
            return
        valores = self.tree_clientes.item(item[0], "values")
        cliente_id = valores[0]
        self.cursor.execute("SELECT * FROM clientes WHERE id=?", (cliente_id,))
        cliente = self.cursor.fetchone()
        if cliente:
            self.abrir_formulario_cliente(cliente)

    def excluir_cliente(self):
        selecionados = self.tree_clientes.selection()
        if not selecionados:
            messagebox.showwarning("Atenção", "Selecione um ou mais clientes para excluir.")
            return

        if not messagebox.askyesno("Confirmação", f"Deseja realmente excluir {len(selecionados)} cliente(s)?"):
            return

        try:
            for item in selecionados:
                valores = self.tree_clientes.item(item, "values")
                cliente_id = valores[0]
                self.cursor.execute("DELETE FROM clientes WHERE id=?", (cliente_id,))
            
            self.conn.commit()
            self.carregar_clientes()
            messagebox.showinfo("Sucesso", f"{len(selecionados)} cliente(s) excluído(s) com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao excluir: {e}")
   
    def criar_aba_consulta_orcamentos(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Consultar Orçamentos")

        search_frame = ttk.LabelFrame(frame, text="Filtros de Busca", padding=10)
        search_frame.pack(fill="x", padx=10, pady=10)  # mantém pack só para o frame em si

        # Número
        ttk.Label(search_frame, text="Número:").pack(side="left", padx=5)
        self.entry_busca_orc = ttk.Entry(search_frame, width=15)
        self.entry_busca_orc.pack(side="left", padx=5)

        # Cliente
        ttk.Label(search_frame, text="Cliente:").pack(side="left", padx=5)
        self.entry_busca_cliente = ttk.Entry(search_frame, width=25)
        self.entry_busca_cliente.pack(side="left", padx=5)  
        # Representante
        ttk.Label(search_frame, text="Representante:").pack(side="left", padx=5)
        self.entry_busca_repr = ttk.Entry(search_frame, width=20)
        self.entry_busca_repr.pack(side="left", padx=5)
        # Status
        ttk.Label(search_frame, text="Status:").pack(side="left", padx=5)
        self.combo_status = ttk.Combobox(search_frame, values=["", "Em Aberto", "Aprovado", "Cancelado", "Rejeitado"], width=15)
        self.combo_status.pack(side="left", padx=5)
        # Datas
        ttk.Label(search_frame, text="Data Inicial (dd/mm/aaaa):").pack(side="left", padx=5)
        self.entry_data_ini = ttk.Entry(search_frame, width=12)
        self.entry_data_ini.pack(side="left", padx=5)
        ttk.Label(search_frame, text="Data Final (dd/mm/aaaa):").pack(side="left", padx=5)
        self.entry_data_fim = ttk.Entry(search_frame, width=12)
        self.entry_data_fim.pack(side="left", padx=5)
        # Botão buscar
        ttk.Button(search_frame, text="Buscar", command=self.buscar_orcamento).pack(side="left", padx=5)

        # Treeview de resultados
        list_frame = ttk.LabelFrame(frame, text="Resultados", padding=10)
        list_frame.pack(fill="both", expand=True, padx=10, pady=10)
        cols = ("Número", "Data", "Cliente", "Total", "Representante", "Status")
        self.tree_orcamentos = ttk.Treeview(list_frame, columns=cols, show="headings", height=12)
        for col in cols:
            self.tree_orcamentos.heading(col, text=col)
            self.tree_orcamentos.column(col, width=140, anchor='center')
        self.tree_orcamentos.pack(fill="both", expand=True)
        # Permitir copiar célula
        self.tree_orcamentos.bind("<Control-c>", self.copiar_celula_treeview)
        self.tree_orcamentos.bind("<Control-Double-1>", self.copiar_celula_treeview, add='+')
        # Configuração de cores só no texto do Status
        self.tree_orcamentos.tag_configure("Em Aberto", foreground="#00EEFF")   # azul
        self.tree_orcamentos.tag_configure("Aprovado", foreground="#00ff5e")    # verde
        self.tree_orcamentos.tag_configure("Rejeitado", foreground="#FF0202")   # vermelho
        self.tree_orcamentos.tag_configure("Cancelado", foreground="#ff5900")   # laranja

        self.tree_orcamentos.bind("<Double-1>", self.visualizar_orcamento)

        self.buscar_orcamento()

    def buscar_orcamento(self):
        numero = self.entry_busca_orc.get().strip()
        cliente = self.entry_busca_cliente.get().strip()
        representante = self.entry_busca_repr.get().strip()
        status = self.combo_status.get().strip()
        data_ini = self.entry_data_ini.get().strip()
        data_fim = self.entry_data_fim.get().strip()

        # Limpa resultados anteriores
        for item in self.tree_orcamentos.get_children():
            self.tree_orcamentos.delete(item)

        # Monta query dinâmica
        query = '''
            SELECT p.numero_pedido, p.data_pedido, c.razao_social, p.valor_total, p.representante, p.status
            FROM pedidos p
            JOIN clientes c ON p.cliente_id = c.id
            WHERE 1=1
        '''
        params = []

        if numero:
            query += " AND p.numero_pedido LIKE ?"
            params.append(f"%{numero}%")
        if cliente:
            query += " AND c.razao_social LIKE ?"
            params.append(f"%{cliente}%")
        if representante:
            query += " AND p.representante LIKE ?"
            params.append(f"%{representante}%")
        if status:
            query += " AND p.status = ?"
            params.append(status)
        if data_ini:
            try:
                dt_ini = datetime.strptime(data_ini, "%d/%m/%Y").strftime("%Y-%m-%d 00:00:00")
                query += " AND p.data_pedido >= ?"
                params.append(dt_ini)
            except:
                messagebox.showwarning("Atenção", "Data inicial inválida! Use dd/mm/aaaa.")
        if data_fim:
            try:
                dt_fim = datetime.strptime(data_fim, "%d/%m/%Y").strftime("%Y-%m-%d 23:59:59")
                query += " AND p.data_pedido <= ?"
                params.append(dt_fim)
            except:
                messagebox.showwarning("Atenção", "Data final inválida! Use dd/mm/aaaa.")

        query += " ORDER BY p.id DESC"

        self.cursor.execute(query, tuple(params))
        rows = self.cursor.fetchall()

        if not rows:
            messagebox.showinfo("Resultado", "Nenhum orçamento encontrado!")
            return

        for row in rows:
            valores = list(row)
            valores[3] = formatar_moeda(valores[3])  # formatar total
            status = valores[5]  # coluna Status

            # insere normalmente
            item_id = self.tree_orcamentos.insert("", "end", values=valores)

            # aplica a cor apenas no campo Status
            self.tree_orcamentos.item(item_id, tags=(status,))

    def salvar_cliente(self):
            try:
                dados = {k: v.get() for k, v in self.cliente_entries.items()}
                if not dados['razao_social'] or not dados['cnpj']:
                    messagebox.showwarning("Atenção", "Razão Social e CNPJ são obrigatórios!")
                    return

                if hasattr(self, 'cliente_edicao_id') and self.cliente_edicao_id:
                    # Atualizar cliente existente
                    self.cursor.execute('''
                        UPDATE clientes
                        SET razao_social=?, cnpj=?, ie=?, endereco=?, cidade=?, estado=?, cep=?, telefone=?, email=?
                        WHERE id=?
                    ''', (dados['razao_social'], dados['cnpj'], dados.get('ie'), dados.get('endereco'),
                        dados.get('cidade'), dados.get('estado'), dados.get('cep'),
                        dados.get('telefone'), dados.get('email'), self.cliente_edicao_id))
                    self.conn.commit()
                    messagebox.showinfo("Sucesso", "Cliente atualizado com sucesso!")
                    self.cliente_edicao_id = None
                else:
                    # Inserir novo cliente
                    self.cursor.execute('''
                        INSERT INTO clientes (razao_social, cnpj, ie, endereco, cidade, estado, cep, telefone, email)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (dados['razao_social'], dados['cnpj'], dados.get('ie'), dados.get('endereco'),
                        dados.get('cidade'), dados.get('estado'), dados.get('cep'),
                        dados.get('telefone'), dados.get('email')))
                    self.conn.commit()
                    messagebox.showinfo("Sucesso", "Cliente cadastrado com sucesso!")

                self.limpar_cliente()
                self.carregar_clientes()
                self.carregar_combos_pedido()
            except sqlite3.IntegrityError:
                messagebox.showerror("Erro", "CNPJ já cadastrado!")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar cliente: {e}")
    
    def limpar_cliente(self):
        for entry in self.cliente_entries.values():
            entry.delete(0, tk.END)
    
    def carregar_clientes(self, filtro=""):
        for item in self.tree_clientes.get_children():
            self.tree_clientes.delete(item)
        if filtro:
            self.cursor.execute('''
                SELECT id, razao_social, cnpj, cidade, telefone 
                FROM clientes 
                WHERE razao_social LIKE ? OR cnpj LIKE ?
            ''', (f'%{filtro}%', f'%{filtro}%'))
        else:
            self.cursor.execute('SELECT razao_social, cnpj, cidade, telefone FROM clientes')
        for row in self.cursor.fetchall():
            self.tree_clientes.insert('', 'end', values=row)

    def editar_cliente(self, event=None):
        item = self.tree_clientes.selection()
        if not item:
            messagebox.showwarning("Atenção", "Selecione um cliente para editar.")
            return
        valores = self.tree_clientes.item(item[0], "values")
        cliente_id = valores[0]
        self.cursor.execute("SELECT * FROM clientes WHERE id=?", (cliente_id,))
        cliente = self.cursor.fetchone()
        if cliente:
            self.abrir_formulario_cliente(cliente)

    # ------------------- Produtos -------------------
    def criar_aba_produtos(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Produtos")
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', padx=10, pady=5)
        ttk.Button(btn_frame, text="Adicionar", bootstyle=SUCCESS,
                command=lambda: self.abrir_formulario_produto()).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Editar", bootstyle=INFO,
                command=self.editar_produto).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Excluir", bootstyle=DANGER,
                command=self.excluir_produto).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Importar Arquivo", bootstyle=WARNING,
                command=lambda: self.importar_dados("produtos")).pack(side='left', padx=5)
        filtro_frame = ttk.Frame(frame)
        filtro_frame.pack(fill='x', padx=10, pady=5)
        ttk.Label(filtro_frame, text="Filtrar por Tipo:").pack(side='left', padx=5)
        self.combo_filtro_tipo = ttk.Combobox(filtro_frame, width=25, state="readonly")
        self.combo_filtro_tipo.pack(side='left', padx=5)
        ttk.Button(filtro_frame, text="Aplicar", command=self.filtrar_produtos_tipo).pack(side='left', padx=5)
        ttk.Button(filtro_frame, text="Limpar", command=lambda: self.carregar_produtos()).pack(side='left', padx=5)
        list_frame = ttk.LabelFrame(frame, text="Produtos Cadastrados", padding=10)
        list_frame.pack(fill='both', expand=True, padx=10, pady=10)
        cols = ('Código', 'Descrição', 'Tipo', 'Origem', 'Valor', 'ICMS%', 'IPI%', 'PIS/COFINS%')
        self.tree_produtos = ttk.Treeview(list_frame, columns=cols, show='headings', height=12)
        for col in cols:
            self.tree_produtos.heading(col, text=col)
            if col in ('ID', 'ICMS%', 'IPI%', 'PIS/COFINS%'):
                self.tree_produtos.column(col, width=80, anchor='center')
            elif col == 'Valor':
                self.tree_produtos.column(col, width=100, anchor='center')
            else:
                self.tree_produtos.column(col, width=160, anchor='center')
        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.tree_produtos.yview)
        self.tree_produtos.configure(yscrollcommand=scrollbar.set)
        self.tree_produtos.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        self.tree_produtos.bind("<Control-c>", self.copiar_celula_treeview)
        self.tree_produtos.bind("<Control-Double-1>", self.copiar_celula_treeview, add='+')
        self.carregar_produtos()

    def salvar_produto(self):
        try:
            dados = {k: v.get() for k, v in self.produto_entries.items()}
            # validar chaves esperadas (usamos normalizar_chave)
            if not dados.get('codigo') or not dados.get('descricao') or not dados.get('valor_unitario'):
                messagebox.showwarning("Atenção", "Código, Descrição e Valor Unitário são obrigatórios!")
                return
            # converter valores
            valor_unit = float(dados.get('valor_unitario') or 0)
            aliq_icms = float(dados.get('icms') or 0)
            aliq_ipi = float(dados.get('ipi') or 0)
            aliq_pis = float(dados.get('pis') or 0)
            aliq_cofins = float(dados.get('cofins') or 0)
            self.cursor.execute('''
                INSERT INTO produtos (codigo, descricao, voltagem, valor_unitario, aliq_icms, aliq_ipi, aliq_pis, aliq_cofins)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (dados.get('codigo'), dados.get('descricao'), dados.get('voltagem'), valor_unit,
                  aliq_icms, aliq_ipi, aliq_pis, aliq_cofins))
            self.conn.commit()
            messagebox.showinfo("Sucesso", "Produto cadastrado com sucesso!")
            self.limpar_produto()
            self.carregar_produtos()
            self.carregar_combos_pedido()
        except sqlite3.IntegrityError:
            messagebox.showerror("Erro", "Código já cadastrado!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar produto: {e}")
    
    def limpar_produto(self):
        for entry in self.produto_entries.values():
            entry.delete(0, tk.END)
    
    def carregar_produtos(self):
        for item in self.tree_produtos.get_children():
            self.tree_produtos.delete(item)

        self.cursor.execute("SELECT DISTINCT tipo FROM produtos WHERE tipo IS NOT NULL AND tipo <> ''")
        tipos = [row[0] for row in self.cursor.fetchall()]
        self.combo_filtro_tipo['values'] = [""] + tipos  # vazio = mostrar todos
        
        # Agora trazemos todas as colunas que o Treeview espera
        self.cursor.execute('''
            SELECT codigo, descricao, tipo, origem_tributacao,
                valor_unitario, aliq_icms, aliq_ipi, aliq_pis
            FROM produtos
        ''')
        # Atualizar combobox com os tipos disponíveis
        for row in self.cursor.fetchall():
            # formatar valor unitário como moeda
            row = list(row)
            row[5] = formatar_moeda(row[5])  # valor_unitario
            self.tree_produtos.insert('', 'end', values=row)
    
    def filtrar_produtos_tipo(self):
        tipo = self.combo_filtro_tipo.get().strip()

        for item in self.tree_produtos.get_children():
            self.tree_produtos.delete(item)

        if not tipo:  # se não selecionar nada, mostra todos
            self.carregar_produtos()
            return

        self.cursor.execute('''
            SELECT id, codigo, descricao, tipo, origem_tributacao,
                valor_unitario, aliq_icms, aliq_ipi, aliq_pis
            FROM produtos
            WHERE tipo = ?
        ''', (tipo,))
        
        for row in self.cursor.fetchall():
            row = list(row)
            row[5] = formatar_moeda(row[5])  # formatar valor unitário
            self.tree_produtos.insert('', 'end', values=row)
    
    def editar_produto(self, event=None):
        item = self.tree_produtos.selection()
        if not item:
            messagebox.showwarning("Atenção", "Selecione um produto para editar.")
            return
        valores = self.tree_produtos.item(item[0], "values")
        produto_id = valores[0]
        self.cursor.execute("SELECT * FROM produtos WHERE id=?", (produto_id,))
        produto = self.cursor.fetchone()
        if produto:
            self.abrir_formulario_produto(produto)
    def excluir_produto(self):
        selecionados = self.tree_produtos.selection()
        if not selecionados:
            messagebox.showwarning("Atenção", "Selecione um ou mais produtos para excluir.")
            return

        if not messagebox.askyesno("Confirmação", f"Deseja realmente excluir {len(selecionados)} produto(s)?"):
            return

        try:
            for item in selecionados:
                valores = self.tree_produtos.item(item, "values")
                produto_id = valores[0]
                self.cursor.execute("DELETE FROM produtos WHERE id=?", (produto_id,))
            
            self.conn.commit()
            self.carregar_produtos()
            messagebox.showinfo("Sucesso", f"{len(selecionados)} produto(s) excluído(s) com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao excluir produto(s): {e}")

    def criar_aba_pedidos(self):
        frame = tb.Frame(self.notebook)
        self.notebook.add(frame, text="Orçamentos")
        # Cabeçalho
        header_frame = tb.Labelframe(frame, text="Cabeçalho do Orçamento", padding=6, bootstyle="primary")
        header_frame.pack(fill='x', padx=10, pady=6)
        tb.Label(header_frame, text="Data do Orçamento:", bootstyle="inverse-primary").grid(row=0, column=0, sticky=W, padx=5)
        self.entry_data_orc = tb.Entry(header_frame, width=15)
        self.entry_data_orc.grid(row=0, column=1, padx=5)
        self.entry_data_orc.insert(0, datetime.now().strftime('%d/%m/%Y'))
        # campo do número do orçamento (só aparece em edição)
        self.label_numero_orc_lbl = tb.Label(header_frame, text="Nº do Orçamento:", bootstyle="inverse-primary")
        self.label_numero_orc = tb.Label(header_frame, text="", bootstyle="secondary")
        tb.Label(header_frame, text="Representante:", bootstyle="inverse-primary").grid(row=0, column=4, sticky=W, padx=5)
        self.entry_representante = tb.Entry(header_frame, width=25)
        self.entry_representante.grid(row=0, column=5, padx=5)
        # 👉 Campo de Status
        # criar mas não mostrar ainda
        self.label_status_orc = tb.Label(header_frame, text="Status:", bootstyle="inverse-primary")
        self.combo_status_orc = tb.Combobox(
            header_frame,
            values=["Em Aberto", "Aprovado", "Cancelado", "Rejeitado"],
            width=15,
            state="readonly"
        )

        # Informações Comerciais
        extra_frame = tb.Labelframe(frame, text="Informações Comerciais", padding=6, bootstyle="info")
        extra_frame.pack(fill='x', padx=10, pady=6)
        tb.Label(extra_frame, text="Condições de Pagamento:").grid(row=0, column=0, sticky=W, padx=5, pady=2)
        self.entry_cond_pag = tb.Entry(extra_frame, width=30)
        self.entry_cond_pag.grid(row=0, column=1, padx=5)
        tb.Label(extra_frame, text="Validade (dias):").grid(row=0, column=2, sticky=W, padx=5, pady=2)
        self.entry_validade = tb.Entry(extra_frame, width=10)
        self.entry_validade.grid(row=0, column=3, padx=5)
        tb.Label(extra_frame, text="Desconto (R$):").grid(row=1, column=0, sticky=W, padx=5, pady=2)
        self.entry_desconto = tb.Entry(extra_frame, width=15)
        self.entry_desconto.grid(row=1, column=1, padx=5)
        tb.Label(extra_frame, text="Observações:").grid(row=2, column=0, sticky=NW, padx=5, pady=2)
        self.text_obs = tk.Text(extra_frame, width=80, height=3)
        self.text_obs.grid(row=2, column=1, columnspan=3, padx=5, pady=2)
        # Seleção cliente / produto
        top_frame = tb.Labelframe(frame, text="Novo Item", padding=6, bootstyle="secondary")
        top_frame.pack(fill='x', padx=10, pady=6)
        tb.Label(top_frame, text="Cliente:").grid(row=0, column=0, sticky=W, padx=5)
        self.combo_cliente = tb.Combobox(top_frame, width=60)
        self.combo_cliente.grid(row=0, column=1, padx=5, columnspan=4)
        tb.Label(top_frame, text="Empresa (emissor):").grid(row=0, column=5, sticky=W, padx=5)
        self.combo_empresa = tb.Combobox(top_frame, width=45)
        self.combo_empresa.grid(row=0, column=6, padx=5, columnspan=2)
        tb.Label(top_frame, text="Produto:").grid(row=1, column=0, sticky=W, padx=5)
        self.combo_produto = tb.Combobox(top_frame, width=60, state='normal')
        self.combo_produto.grid(row=1, column=1, padx=5, columnspan=3)
        tb.Label(top_frame, text="Quantidade:").grid(row=1, column=4, sticky=W, padx=5)
        self.entry_qtd = tb.Entry(top_frame, width=8)
        self.entry_qtd.grid(row=1, column=5, padx=5)
        tb.Button(top_frame, text="Adicionar Item", command=self.adicionar_item_pedido, bootstyle="success").grid(row=1, column=6, padx=5)
        tb.Button(top_frame, text="Remover Item", command=self.remover_item, bootstyle="danger").grid(row=1, column=7, padx=5)
        items_frame = tb.Labelframe(frame, text="Itens do Orçamento", padding=6, bootstyle="info")
        items_frame.pack(fill='both', expand=True, padx=10, pady=6)
        cols = ('Produto', 'Qtd', 'Valor Unit.', 'Total')
        self.tree_pedido_items = tb.Treeview(
            items_frame, columns=cols, show='headings', height=9, bootstyle="info"
        )

        for col in cols:
            self.tree_pedido_items.heading(col, text=col)
            # Todas as colunas centralizadas e com largura proporcional
            if col == 'Produto':
                self.tree_pedido_items.column(col, width=300, anchor='center')
            elif col == 'Qtd':
                self.tree_pedido_items.column(col, width=80, anchor='center')
            elif col == 'Valor Unit.':
                self.tree_pedido_items.column(col, width=120, anchor='center')
            elif col == 'Total':
                self.tree_pedido_items.column(col, width=120, anchor='center')

        self.tree_pedido_items.pack(fill='both', expand=True)

        # Totais + ações
        totals_frame = tb.Labelframe(frame, text="Totais & Ações", padding=6, bootstyle="warning")
        totals_frame.pack(fill='x', padx=10, pady=6)
        self.label_subtotal = tb.Label(totals_frame, text="Subtotal: R$ 0,00", bootstyle="secondary")
        self.label_subtotal.grid(row=0, column=0, padx=8)
        self.label_impostos = tb.Label(totals_frame, text="Impostos: R$ 0,00", bootstyle="secondary")
        self.label_impostos.grid(row=0, column=1, padx=8)
        self.label_total = tb.Label(totals_frame, text="TOTAL: R$ 0,00",
                                    font=('Segoe UI', 11, 'bold'), bootstyle="success")
        self.label_total.grid(row=0, column=2, padx=8)
        self.btn_finalizar_pedido = tb.Button(
            totals_frame, text="Salvar Orçamento",
            command=self.finalizar_pedido, bootstyle="success"
        )
        self.btn_finalizar_pedido.grid(row=0, column=3, padx=8)
        tb.Button(
            totals_frame, text="Novo Orçamento",
            command=self.limpar_pedido, bootstyle="secondary"
        ).grid(row=0, column=4, padx=8)

        tb.Button(
            totals_frame, text="Exportar p/ PDF",
            command=lambda: self.gerar_pdf_orcamento(self.label_numero_orc.cget("text")),
            bootstyle="danger-outline"
        ).grid(row=0, column=6, padx=8)

        # carregar clientes e produtos nos combos
        self.carregar_combos_pedido()


    def carregar_orcamento_para_edicao(self, numero_pedido):
        """
        Carrega um orçamento salvo para edição em uma nova aba separada.
        """
        # Remove aba antiga de edição se já existir
        for i, tab_id in enumerate(self.notebook.tabs()):
            if "Editar Orçamento" in self.notebook.tab(tab_id, "text"):
                self.notebook.forget(tab_id)
                break

        # Cria nova aba
        aba_editar = tb.Frame(self.notebook)
        self.notebook.add(aba_editar, text=f"Editar Orçamento ({numero_pedido})")
        self.notebook.select(aba_editar)

        # Buscar dados principais
        self.cursor.execute('''
            SELECT p.data_pedido, p.cliente_id, p.valor_produtos, p.valor_icms, p.valor_ipi,
                p.valor_pis, p.valor_cofins, p.valor_total, p.representante,
                p.condicoes_pagamento, p.desconto, p.observacoes, p.validade, p.status,
                p.empresa_id
            FROM pedidos p
            WHERE p.numero_pedido = ?
        ''', (numero_pedido,))
        pedido = self.cursor.fetchone()
        if not pedido:
            messagebox.showerror("Erro", "Orçamento não encontrado.")
            return

        (data_pedido, cliente_id, subtotal, icms, ipi, pis, cofins, total,
        representante, cond_pag, desconto, observacoes, validade, status, empresa_id) = pedido

        # --- Cabeçalho ---
        header = tb.Labelframe(aba_editar, text="Cabeçalho do Orçamento", bootstyle="primary", padding=8)
        header.pack(fill='x', padx=10, pady=10)

        tb.Label(header, text="Nº Orçamento:").grid(row=0, column=0, sticky="w", padx=5)
        tb.Label(header, text=numero_pedido, bootstyle="secondary").grid(row=0, column=1, padx=5)

        tb.Label(header, text="Data:").grid(row=0, column=2, sticky="w", padx=5)
        entry_data = tb.Entry(header, width=15)
        entry_data.grid(row=0, column=3, padx=5)
        try:
            entry_data.insert(0, datetime.strptime(data_pedido, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y"))
        except:
            entry_data.insert(0, data_pedido)

        tb.Label(header, text="Status:").grid(row=0, column=4, sticky="w", padx=5)
        combo_status = tb.Combobox(header, values=["Em Aberto", "Aprovado", "Cancelado", "Rejeitado"], width=15)
        combo_status.set(status or "Em Aberto")
        combo_status.grid(row=0, column=5, padx=5)

        tb.Label(header, text="Representante:").grid(row=1, column=0, sticky="w", padx=5)
        entry_repr = tb.Entry(header, width=30)
        entry_repr.grid(row=1, column=1, padx=5, columnspan=2)
        entry_repr.insert(0, representante or "")

        tb.Label(header, text="Cliente:").grid(row=1, column=3, sticky="w", padx=5)
        combo_cliente = tb.Combobox(header, width=60)
        combo_cliente.grid(row=1, column=4, columnspan=2, padx=5)
        self.cursor.execute("SELECT id, razao_social FROM clientes")
        clientes = [f"{cid} - {nome}" for cid, nome in self.cursor.fetchall()]
        combo_cliente['values'] = clientes
        try:
            self.cursor.execute("SELECT razao_social FROM clientes WHERE id=?", (cliente_id,))
            nome_cliente = self.cursor.fetchone()[0]
            combo_cliente.set(f"{cliente_id} - {nome_cliente}")
        except:
            pass

        # Empresa emissora
        tb.Label(header, text="Empresa (emissor):").grid(row=2, column=0, sticky="w", padx=5)
        combo_empresa = tb.Combobox(header, width=60)
        combo_empresa.grid(row=2, column=1, columnspan=3, padx=5)
        self.cursor.execute("SELECT id, nome FROM empresas")
        empresas = [f"{eid} - {nome}" for eid, nome in self.cursor.fetchall()]
        combo_empresa['values'] = empresas
        if empresa_id:
            try:
                self.cursor.execute("SELECT nome FROM empresas WHERE id=?", (empresa_id,))
                emp_nome = self.cursor.fetchone()[0]
                combo_empresa.set(f"{empresa_id} - {emp_nome}")
            except:
                pass

        # --- Itens do orçamento ---
        frame_itens = tb.Labelframe(aba_editar, text="Itens do Orçamento", bootstyle="info", padding=8)
        frame_itens.pack(fill="both", expand=True, padx=10, pady=5)
        cols = ("Produto", "Qtd", "Valor Unit.", "Total")
        tree_itens = tb.Treeview(frame_itens, columns=cols, show="headings", bootstyle="info")
        for col in cols:
            tree_itens.heading(col, text=col)
            tree_itens.column(col, width=160, anchor="center")
        tree_itens.pack(fill="both", expand=True)

        # Carregar itens do banco
        self.cursor.execute('''
            SELECT pr.id, pr.codigo, pr.descricao, pi.qtd, pi.valor_unitario
            FROM pedido_itens pi
            JOIN produtos pr ON pi.produto_id = pr.id
            WHERE pi.numero_pedido = ?
        ''', (numero_pedido,))
        itens = self.cursor.fetchall()

        itens_temp = []
        for prod_id, codigo, descricao, qtd, valor_unit in itens:
            total_item = qtd * valor_unit
            itens_temp.append({'produto_id': prod_id, 'codigo': codigo, 'descricao': descricao,
                            'qtd': qtd, 'valor': valor_unit})
            tree_itens.insert('', 'end', values=(
                f"{codigo} - {descricao}", qtd, formatar_moeda(valor_unit), formatar_moeda(total_item)
            ))

        # --- Totais e ações ---
        frame_totais = tb.Labelframe(aba_editar, text="Totais & Ações", bootstyle="warning", padding=8)
        frame_totais.pack(fill='x', padx=10, pady=5)

        label_total = tb.Label(frame_totais, text=f"TOTAL: {formatar_moeda(total)}",
                            font=('Segoe UI', 11, 'bold'), bootstyle="success")
        label_total.pack(side="left", padx=8)

        def atualizar_orcamento():
            try:
                cliente_id_sel = int(combo_cliente.get().split(" - ")[0])
                status_sel = combo_status.get()
                repr_sel = entry_repr.get()
                cond_pag = cond_pag_entry.get()
                validade = entry_validade.get()
                desconto_val = float(entry_desc.get() or 0)
                obs_texto = txt_obs.get("1.0", tk.END).strip()

                self.cursor.execute('''
                    UPDATE pedidos
                    SET cliente_id=?, representante=?, condicoes_pagamento=?, desconto=?, observacoes=?, validade=?, status=?
                    WHERE numero_pedido=?
                ''', (cliente_id_sel, repr_sel, cond_pag, desconto_val, obs_texto, validade, status_sel, numero_pedido))
                self.conn.commit()
                messagebox.showinfo("Sucesso", f"Orçamento {numero_pedido} atualizado com sucesso!")
                self.buscar_orcamento()
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao atualizar: {e}")

        tb.Button(frame_totais, text="Atualizar Orçamento", bootstyle="success", command=atualizar_orcamento).pack(side="right", padx=8)
        tb.Button(frame_totais, text="Gerar PDF", bootstyle="danger-outline",
                command=lambda: self.gerar_pdf_orcamento(numero_pedido)).pack(side="right", padx=8)

        # --- Informações adicionais ---
        frame_extra = tb.Labelframe(aba_editar, text="Informações Comerciais", bootstyle="secondary", padding=8)
        frame_extra.pack(fill='x', padx=10, pady=5)

        tb.Label(frame_extra, text="Condições de Pagamento:").grid(row=0, column=0, sticky="w", padx=5)
        cond_pag_entry = tb.Entry(frame_extra, width=40)
        cond_pag_entry.grid(row=0, column=1, padx=5)
        cond_pag_entry.insert(0, cond_pag or "")

        tb.Label(frame_extra, text="Validade (dias):").grid(row=0, column=2, sticky="w", padx=5)
        entry_validade = tb.Entry(frame_extra, width=10)
        entry_validade.grid(row=0, column=3, padx=5)
        entry_validade.insert(0, validade or "")

        tb.Label(frame_extra, text="Desconto (R$):").grid(row=1, column=0, sticky="w", padx=5)
        entry_desc = tb.Entry(frame_extra, width=15)
        entry_desc.grid(row=1, column=1, padx=5)
        entry_desc.insert(0, str(desconto or 0))

        tb.Label(frame_extra, text="Observações:").grid(row=2, column=0, sticky="nw", padx=5)
        txt_obs = tk.Text(frame_extra, width=80, height=3)
        txt_obs.grid(row=2, column=1, columnspan=3, padx=5)
        txt_obs.insert("1.0", observacoes or "")


    def remover_item(self):
        selecionado = self.tree_pedido_items.selection()
        if not selecionado:
            messagebox.showwarning("Atenção", "Selecione um item para remover.")
            return

        for item_id in selecionado:
            valores = self.tree_pedido_items.item(item_id, "values")
            if valores:
                descricao = valores[0]  
                for i, item in enumerate(self.itens_pedido_temp):
                    if f"{item['codigo']} - {item['descricao']}" == descricao:
                        del self.itens_pedido_temp[i]
                        break
           
            self.tree_pedido_items.delete(item_id)

        self.atualizar_totais() 

    def carregar_combos_pedido(self):
        self.cursor.execute('SELECT id, razao_social FROM clientes')
        clientes = [f"{row[0]} - {row[1]}" for row in self.cursor.fetchall()]
        self.combo_cliente['values'] = clientes     
        self.cursor.execute('SELECT id, codigo, descricao FROM produtos')
        produtos = [f"{row[0]} - {row[1]} - {row[2]}" for row in self.cursor.fetchall()]
        self.combo_produto['values'] = produtos
        # carregar empresas
        self.cursor.execute('SELECT id, nome FROM empresas')
        empresas = [f"{row[0]} - {row[1]}" for row in self.cursor.fetchall()]
        self.combo_empresa['values'] = empresas

    def filtrar_clientes(self, event=None):
        texto = self.combo_cliente.get().lower()
        self.cursor.execute('SELECT id, razao_social FROM clientes')
        todos = [f"{row[0]} - {row[1]}" for row in self.cursor.fetchall()]

        if texto:
            filtrados = [c for c in todos if texto in c.lower()]
        else:
            filtrados = todos

        self.combo_cliente['values'] = filtrados
        self.combo_cliente.event_generate('<Down>')  # abre a lista automaticamente

    def adicionar_item_pedido(self):
        try:
            produto_str = self.combo_produto.get()
            if not produto_str:
                messagebox.showwarning("Atenção", "Selecione um produto!")
                return

            produto_id = int(produto_str.split(" - ")[0])
            qtd = float(self.entry_qtd.get() or 1)

            self.cursor.execute("SELECT codigo, descricao, valor_unitario FROM produtos WHERE id = ?", (produto_id,))
            produto = self.cursor.fetchone()
            if not produto:
                messagebox.showerror("Erro", "Produto não encontrado.")
                return

            codigo, descricao, preco_venda = produto

            # Verifica se já existe o mesmo produto na lista
            encontrado = False
            for item in self.itens_pedido_temp:
                if item["produto_id"] == produto_id:
                    # Atualiza quantidade e recalcula
                    item["qtd"] += qtd
                    encontrado = True
                    break

            if not encontrado:
                # adiciona novo item
                self.itens_pedido_temp.append({
                    "produto_id": produto_id,
                    "codigo": codigo,
                    "descricao": descricao,
                    "qtd": qtd,
                    "valor": float(preco_venda or 0)
                })

            # Atualiza visualização na treeview
            self.tree_pedido_items.delete(*self.tree_pedido_items.get_children())
            for item in self.itens_pedido_temp:
                total_item = item["qtd"] * item["valor"]
                self.tree_pedido_items.insert(
                    "", "end",
                    values=(f"{item['codigo']} - {item['descricao']}",
                            item["qtd"],
                            formatar_moeda(item["valor"]),
                            formatar_moeda(total_item))
                )

            self.atualizar_totais()
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao adicionar item: {e}")

    def atualizar_totais(self):
        subtotal = sum(item['qtd'] * item['valor'] for item in self.itens_pedido_temp)
        total_icms = 0
        total_ipi = 0
        total_pis = 0
        total_cofins = 0

        for item in self.itens_pedido_temp:
            self.cursor.execute('''
                SELECT aliq_icms, aliq_ipi, aliq_pis, aliq_cofins
                FROM produtos WHERE id = ?
            ''', (item['produto_id'],))
            aliq_icms, aliq_ipi, aliq_pis, aliq_cofins = self.cursor.fetchone()
            
            base = item['qtd'] * item['valor']
            total_icms   += base * (aliq_icms or 0) / 100
            total_ipi    += base * (aliq_ipi or 0) / 100
            total_pis    += base * (aliq_pis or 0) / 100
            total_cofins += base * (aliq_cofins or 0) / 100

        total_impostos = total_icms + total_ipi + total_pis + total_cofins
        desconto = float(self.entry_desconto.get() or 0)
        total = subtotal + total_impostos - desconto

        self.label_subtotal.config(text=f"Subtotal: {formatar_moeda(subtotal)}")
        self.label_impostos.config(text=f"Impostos: {formatar_moeda(total_impostos)}")
        self.label_total.config(text=f"TOTAL: {formatar_moeda(total)}")
    
    def visualizar_cliente(self, event):
        item = self.tree_clientes.selection()
        if not item:
            return
        
        valores = self.tree_clientes.item(item[0], "values")
        cliente_id = valores[0]  
        self.cursor.execute("SELECT * FROM clientes WHERE id=?", (cliente_id,))
        cliente = self.cursor.fetchone()

        if not cliente:
            messagebox.showerror("Erro", "Cliente não encontrado.")
            return

        keys = ["ID", "Razão Social", "CNPJ", "IE", "Endereço", "Cidade", "Estado", "CEP", "Telefone", "Email"]

        top = tk.Toplevel(self.root)
        top.title(f"Cliente - {cliente[1]}")
        top.geometry("600x400")

        frame_info = ttk.LabelFrame(top, text="Dados do Cliente", padding=10)
        frame_info.pack(fill="both", expand=True, padx=10, pady=10)

        # Mostrar cada campo em label
        for i, (k, v) in enumerate(zip(keys, cliente)):
            tk.Label(frame_info, text=f"{k}:", font=("Arial", 10, "bold")).grid(row=i, column=0, sticky="w", padx=5, pady=2)
            tk.Label(frame_info, text=v if v else "Não informado").grid(row=i, column=1, sticky="w", padx=5, pady=2)

        # Botões de ação
        frame_botoes = ttk.Frame(top)
        frame_botoes.pack(pady=10)

        def acao_editar():
            """Chama o método editar_cliente para preencher o formulário da aba."""
            self.tree_clientes.selection_set(item)   
            self.editar_cliente(None)                
            top.destroy()                            # fecha a janelinha

        ttk.Button(frame_botoes, text="Editar Cliente", command=acao_editar).pack(side="left", padx=5)
        ttk.Button(frame_botoes, text="Fechar", command=top.destroy).pack(side="left", padx=5)
    
    def visualizar_orcamento(self, event):
        item = self.tree_orcamentos.selection()
        if not item:
            return
        
        valores = self.tree_orcamentos.item(item[0], "values")
        numero_pedido = valores[0]  # primeira coluna = Número do Orçamento
        
        # Nova janela
        top = tk.Toplevel(self.root)
        top.title(f"Orçamento {numero_pedido}")
        top.geometry("850x600")

        # Buscar dados principais do pedido
        self.cursor.execute('''
            SELECT p.data_pedido, c.razao_social, c.cnpj, c.endereco, c.cidade, c.estado,
                p.valor_total, p.representante, p.condicoes_pagamento, p.observacoes, p.validade
            FROM pedidos p
            JOIN clientes c ON p.cliente_id = c.id
            WHERE p.numero_pedido = ?
        ''', (numero_pedido,))
        pedido = self.cursor.fetchone()

        if not pedido:
            messagebox.showerror("Erro", "Orçamento não encontrado.")
            top.destroy()
            return

        data, cliente, cnpj, endereco, cidade, estado, total, representante, cond_pag, obs, validade = pedido

        # Cabeçalho
        frame_info = ttk.LabelFrame(top, text="Dados do Orçamento", padding=10)
        frame_info.pack(fill="x", padx=10, pady=10)
        partes_endereco = []
        if endereco:
            partes_endereco.append(endereco)
        if cidade:
            partes_endereco.append(cidade)
        if estado:
            partes_endereco.append(estado)

        texto_endereco = " - ".join(partes_endereco) if partes_endereco else "Não informado"
        tk.Label(frame_info, text=f"Data: {data}").grid(row=0, column=0, sticky="w", padx=5)
        tk.Label(frame_info, text=f"Cliente: {cliente}").grid(row=1, column=0, sticky="w", padx=5)
        tk.Label(frame_info, text=f"CNPJ: {cnpj}").grid(row=1, column=1, sticky="w", padx=5)
        tk.Label(frame_info, text=f"Endereço: {texto_endereco}").grid(row=2, column=0, columnspan=2, sticky="w", padx=5)
        tk.Label(frame_info, text=f"Representante: {representante}").grid(row=3, column=0, sticky="w", padx=5)
        tk.Label(frame_info, text=f"Condições de Pagamento: {cond_pag}").grid(row=4, column=0, sticky="w", padx=5)
        tk.Label(frame_info, text=f"Validade: {validade} dias").grid(row=4, column=1, sticky="w", padx=5)
        tk.Label(frame_info, text=f"Observações: {obs}").grid(row=5, column=0, columnspan=2, sticky="w", padx=5)
        tk.Label(frame_info, text=f"TOTAL: {formatar_moeda(total)}", font=("Arial", 11, "bold")).grid(row=6, column=0, sticky="w", padx=5, pady=5)

        # Itens do orçamento
        frame_itens = ttk.LabelFrame(top, text="Itens do Orçamento", padding=10)
        frame_itens.pack(fill="both", expand=True, padx=10, pady=10)

        cols = ("Código", "Descrição", "Qtd", "Valor Unit.", "Total")
        tree_itens = ttk.Treeview(frame_itens, columns=cols, show="headings", height=10)
        for col in cols:
            tree_itens.heading(col, text=col, anchor="center")
            tree_itens.column(col, width=130, anchor="center")
        tree_itens.pack(fill="both", expand=True)

        # Carregar itens do pedido
        self.cursor.execute('''
            SELECT pr.codigo, pr.descricao, pi.qtd, pi.valor_unitario
            FROM pedido_itens pi
            JOIN produtos pr ON pi.produto_id = pr.id
            WHERE pi.numero_pedido = ?
        ''', (numero_pedido,))
        itens = self.cursor.fetchall()

        for cod, desc, qtd, valor_unit in itens:
            tree_itens.insert("", "end", values=(
                cod, desc, qtd, formatar_moeda(valor_unit), formatar_moeda(qtd * valor_unit)
            ))

        # Rodapé com botões
        frame_botoes = ttk.Frame(top)
        frame_botoes.pack(fill="x", pady=10)
        ttk.Button(frame_botoes, text="Exportar p/ Excel", bootstyle=DANGER,command=self.exportar_excel_orcamento).pack(side="left", padx=5)
        ttk.Button(frame_botoes, bootstyle=INFO, text="Editar em Nova Aba", 
        command=lambda: (top.destroy(), self.carregar_orcamento_para_edicao(numero_pedido))).pack(side="left", padx=5)
        ttk.Button(frame_botoes, text="Fechar", command=top.destroy).pack(side="right", padx=5)
        
    def limpar_pedido(self):
        for item in self.tree_pedido_items.get_children():
            self.tree_pedido_items.delete(item)
            self.itens_pedido_temp = []
            self.atualizar_totais()
            self.combo_cliente.set('')
            self.combo_produto.set('')
            self.entry_qtd.delete(0, tk.END)

        try:
            self.label_status_orc.grid_forget()
            self.combo_status_orc.grid_forget()
        except:
            pass
        try:
            self.label_numero_orc_lbl.grid_forget()
            self.label_numero_orc.grid_forget()
        except:
            pass

    def finalizar_pedido(self):
        if not self.combo_cliente.get() or not self.itens_pedido_temp:
            messagebox.showwarning("Atenção", "Selecione um cliente e adicione itens!")
            return
        try:
            cliente_id = int(self.combo_cliente.get().split(' - ')[0])
            # se estiver em edição, usamos o numero existente
            if hasattr(self, 'edicao_numero_pedido') and self.edicao_numero_pedido:
                numero_pedido = self.edicao_numero_pedido
            else:
                self.cursor.execute("SELECT COUNT(*) FROM pedidos")
                total_registros = self.cursor.fetchone()[0] or 0
                data = datetime.now().strftime("%Y%m%d")
                numero_pedido = f"ORC-{data}-{total_registros+1:03d}"
                self.label_numero_orc.config(text=numero_pedido)

            data_pedido = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # calcular totais
            subtotal, total_icms, total_ipi, total_pis, total_cofins, total = self.calcular_totais(self.itens_pedido_temp)

            representante = self.entry_representante.get()
            cond_pag = self.entry_cond_pag.get()
            validade = self.entry_validade.get()
            desconto = float(self.entry_desconto.get() or 0)
            observacoes = self.text_obs.get("1.0", tk.END).strip()

            total_final = total - desconto

            empresa_id = None
            if self.combo_empresa.get():
                try:
                    empresa_id = int(self.combo_empresa.get().split(' - ')[0])
                except:
                    empresa_id = None

            if hasattr(self, 'edicao_numero_pedido') and self.edicao_numero_pedido:
                # === Atualizar pedido existente ===
                status = self.combo_status_orc.get() or "Em Aberto"  # pega do combobox

                self.cursor.execute('''
                    UPDATE pedidos
                    SET data_pedido=?, cliente_id=?, valor_produtos=?, valor_icms=?, valor_ipi=?, valor_pis=?, valor_cofins=?, valor_total=?,
                        representante=?, condicoes_pagamento=?, desconto=?, status=?, observacoes=?, validade=?, empresa_id=?
                    WHERE numero_pedido=?
                ''', (data_pedido, cliente_id, subtotal, total_icms, total_ipi, total_pis, total_cofins, total_final,
                    representante, cond_pag, desconto, status, observacoes, validade, numero_pedido, empresa_id))

                # remover itens antigos e inserir os novos
                self.cursor.execute('DELETE FROM pedido_itens WHERE numero_pedido = ?', (numero_pedido,))
                for item in self.itens_pedido_temp:
                    self.cursor.execute('''
                        INSERT INTO pedido_itens (numero_pedido, produto_id, qtd, valor_unitario)
                        VALUES (?, ?, ?, ?)
                    ''', (numero_pedido, item['produto_id'], item['qtd'], item['valor']))

                self.conn.commit()
                messagebox.showinfo("Sucesso", f"Orçamento {numero_pedido} atualizado com sucesso!")
                # sair do modo edição
                self.edicao_numero_pedido = None
                try:
                    self.btn_finalizar_pedido.config(text="Salvar Orçamento")
                except:
                    pass
            else:
                # === Inserir novo pedido ===
                self.cursor.execute('''
                    INSERT INTO pedidos 
                    (numero_pedido, data_pedido, cliente_id, valor_produtos, valor_icms, valor_ipi, valor_pis, valor_cofins, valor_total, representante,
                    condicoes_pagamento, desconto, status, observacoes, validade, empresa_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (numero_pedido, data_pedido, cliente_id, subtotal, total_icms, total_ipi, total_pis, total_cofins, total_final,
                    representante, cond_pag, desconto, "Em Aberto", observacoes, validade, empresa_id))

                for item in self.itens_pedido_temp:
                    self.cursor.execute('''
                        INSERT INTO pedido_itens (numero_pedido, produto_id, qtd, valor_unitario)
                        VALUES (?, ?, ?, ?)
                    ''', (numero_pedido, item['produto_id'], item['qtd'], item['valor']))

                self.conn.commit()
                messagebox.showinfo("Sucesso", f"Orçamento {numero_pedido} salvo com sucesso!")

            # limpar itens temporários e UI
            self.itens_pedido_temp.clear()
            self.tree_pedido_items.delete(*self.tree_pedido_items.get_children())
            self.carregar_combos_pedido()
            try:
                # recarregar a lista de orçamentos (se existir)
                self.buscar_orcamento()
            except:
                pass

        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao salvar/atualizar orçamento: {e}")

    def abrir_formulario_produto(self, produto=None):
        top = tk.Toplevel(self.root)
        top.title("Cadastro de Produto")
        top.geometry("700x400")

        labels = [
            'Código:', 'Descrição:', 'Valor Unitário:',
            'Tipo:', 'Origem Tributação:', 'Voltagem:',
            'ICMS (%):', 'IPI (%):', 'PIS (%):', 'COFINS (%):'
        ]

        entries = {}
        for i, label in enumerate(labels):
            ttk.Label(top, text=label).grid(row=i//2, column=(i%2)*2, sticky='w', padx=5, pady=5)
            entry = ttk.Entry(top, width=28)
            entry.grid(row=i//2, column=(i%2)*2+1, padx=5, pady=5)
            chave = normalizar_chave(label)
            entries[chave] = entry

        # se for edição, preencher
        if produto:
            keys = ['id','codigo','descricao','valor_unitario','tipo','origem_tributacao','voltagem',
                    'aliq_icms','aliq_ipi','aliq_pis','aliq_cofins']
            for k, v in zip(keys, produto):
                if k in entries:
                    entries[k].insert(0, v or "")

            def salvar():
                dados = {k: v.get().strip() for k,v in entries.items()}
                if not dados['codigo']:
                    messagebox.showerror("Erro", "O campo CÓDIGO do produto está vazio.")
                    return
                if not dados['descricao']:
                    messagebox.showerror("Erro", "O campo DESCRIÇÃO do produto está vazio.")
                    return
                if not dados['valor_unitario']:
                    messagebox.showerror("Erro", "O campo VALOR UNITÁRIO está vazio.")
                    return

                try:
                    if produto:  # atualização
                        self.cursor.execute('''
                            UPDATE produtos
                            SET codigo=?, descricao=?, valor_unitario=?, tipo=?, origem_tributacao=?, voltagem=?,
                                aliq_icms=?, aliq_ipi=?, aliq_pis=?, aliq_cofins=?
                            WHERE id=?
                        ''', (dados['codigo'], dados['descricao'], float(dados['valor_unitario'] or 0),
                            dados.get('tipo'), dados.get('origem_tributacao'), dados.get('voltagem'),
                            float(dados.get('icms') or 0), float(dados.get('ipi') or 0),
                            float(dados.get('pis') or 0), float(dados.get('cofins') or 0),
                            produto[0]))
                    else:  # novo
                        self.cursor.execute('''
                            INSERT INTO produtos (codigo, descricao, valor_unitario, tipo, origem_tributacao, voltagem,
                                                aliq_icms, aliq_ipi, aliq_pis, aliq_cofins)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (dados['codigo'], dados['descricao'], float(dados['valor_unitario'] or 0),
                            dados.get('tipo'), dados.get('origem_tributacao'), dados.get('voltagem'),
                            float(dados.get('icms') or 0), float(dados.get('ipi') or 0),
                            float(dados.get('pis') or 0), float(dados.get('cofins') or 0)))
                    self.conn.commit()
                    self.carregar_produtos()
                    top.destroy()
                except sqlite3.IntegrityError:
                    messagebox.showerror("Erro", "Código já cadastrado!")   
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao salvar produto: {e}")

            ttk.Button(top, text="Salvar", command=salvar).grid(row=6, column=0, columnspan=2, pady=10)

    # ------------------- Export PDF -------------------
    def gerar_pdf_orcamento(self, numero_pedido=None):
        try:
            if not numero_pedido:
                selecionado = self.tree_orcamentos.selection()
                if not selecionado:
                    messagebox.showwarning("Aviso", "Selecione um orçamento para gerar o PDF.")
                    return
                numero_pedido = self.tree_orcamentos.item(selecionado, 'values')[0]
            self.cursor.execute('''
                SELECT p.numero_pedido, p.data_pedido, c.razao_social, c.cnpj, c.endereco, c.cidade, c.estado,
                    p.valor_produtos, p.valor_icms, p.valor_ipi, p.valor_pis, p.valor_cofins, 
                    p.valor_total, p.representante, p.condicoes_pagamento, p.desconto, p.status,
                    p.observacoes, p.validade, p.empresa_id
                FROM pedidos p
                LEFT JOIN clientes c ON p.cliente_id = c.id
                WHERE p.numero_pedido = ?
            ''', (numero_pedido,))
            pedido = self.cursor.fetchone()

            if not pedido:
                messagebox.showerror("Erro", "Orçamento não encontrado.")
                return

            (num, data, cliente_nome, cliente_cnpj, cliente_end, cliente_cid, cliente_est,
            subtotal, icms, ipi, pis, cofins, total, representante, cond_pag, desconto,
            status, observacoes, validade, empresa_id) = pedido

            # === Buscar dados da empresa emissora ===
            if empresa_id:
                self.cursor.execute("SELECT nome, cnpj, endereco, cidade, estado, telefone, email, caminho_logo FROM empresas WHERE id = ?", (empresa_id,))
                emp = self.cursor.fetchone()
            else:
                emp = None

            if emp:
                nome_emp, cnpj_emp, end_emp, cid_emp, est_emp, tel_emp, email_emp, logo_emp = emp
            else:
                nome_emp = "Eletrofrio Refrigeração Ltda."
                cnpj_emp = "76.498.179/0001-10"
                end_emp = "Rua João Chede, 1599 – CIC, Curitiba/PR, CEP 81170-220"
                cid_emp = "Curitiba"
                est_emp = "PR"
                tel_emp = "(41) 2105-6000"
                email_emp = "marketing@eletrofrio.com.br"
                logo_emp = "logo.png"

            # === Buscar itens do orçamento ===
            self.cursor.execute('''
                SELECT pr.codigo, pr.descricao, i.qtd, i.valor_unitario
                FROM pedido_itens i
                LEFT JOIN produtos pr ON i.produto_id = pr.id
                WHERE i.numero_pedido = ?
            ''', (numero_pedido,))
            itens = self.cursor.fetchall()
            # === Caminho do arquivo ===
            nome_pdf = f"orcamento-{cliente_nome.replace(' ', '_')}-{datetime.now().strftime('%d-%m-%y')}.pdf"
            caminho_pdf = os.path.join(os.getcwd(), nome_pdf)
            # === Criar documento ===
            doc = SimpleDocTemplate(caminho_pdf, pagesize=A4, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=30)
            story = []
            styles = getSampleStyleSheet()
            estilo_normal = styles["Normal"]
            dados_cabecalho = []
            # Logo (se existir)
            logo_path = logo_emp if os.path.exists(logo_emp) else None
            if logo_path:
                img = PDFImage(logo_path, width=120, height=50)
            else:
                img = Paragraph("<b>Sem Logo</b>", estilo_normal)

            info_empresa = f"""
            <b>{nome_emp}</b><br/>
            CNPJ: {cnpj_emp}<br/>
            {end_emp}, {cid_emp} - {est_emp}<br/>
            {email_emp} | {tel_emp}
            """
            dados_cabecalho.append([img, Paragraph(info_empresa, estilo_normal)])

            tabela_cab = Table(dados_cabecalho, colWidths=[130, 350])
            tabela_cab.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]))
            story.append(tabela_cab)
            story.append(Spacer(1, 10))
            story.append(Table([[""]], colWidths=[500], style=[("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.grey)]))
            story.append(Spacer(1, 10))

            # === Dados do orçamento ===
            dados_orc = f"""
            <b>ORÇAMENTO Nº:</b> {num} &nbsp;&nbsp;&nbsp;&nbsp; <b>Data:</b> {data}<br/>
            <b>Cliente:</b> {cliente_nome}<br/>
            <b>CNPJ:</b> {cliente_cnpj}<br/>
            <b>Endereço:</b> {cliente_end}, {cliente_cid} - {cliente_est}<br/>
            <b>Representante:</b> {representante} &nbsp;&nbsp;&nbsp;&nbsp; <b>Status:</b> {status}<br/>
            <b>Validade:</b> {validade or '-'} dias &nbsp;&nbsp;&nbsp;&nbsp; <b>Pagamento:</b> {cond_pag or '-'}

            """
            story.append(Paragraph(dados_orc, estilo_normal))
            story.append(Spacer(1, 10))

            # === Tabela de itens ===
            cabecalho = ["Código", "Descrição", "Qtd", "V. Unitário", "Total"]
            linhas = [cabecalho]
            for cod, desc, qtd, valor in itens:
                total_item = float(qtd) * float(valor)
                linhas.append([
                    cod,
                    Paragraph(desc, estilo_normal),
                    f"{qtd:.0f}",
                    f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                    f"R$ {total_item:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                ])

            tabela = Table(linhas, colWidths=[70, 240, 50, 80, 80])
            tabela.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]))
            story.append(tabela)
            story.append(Spacer(1, 15))

            # === Totais ===
            dados_totais = [
                ["Subtotal:", f"R$ {subtotal:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")],
                ["ICMS:", f"R$ {icms:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")],
                ["IPI:", f"R$ {ipi:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")],
                ["PIS:", f"R$ {pis:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")],
                ["COFINS:", f"R$ {cofins:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")],
                ["Desconto:", f"R$ {desconto:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")],
                ["", ""],
                ["Total Geral:", f"R$ {total:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")],
            ]
            tabela_totais = Table(dados_totais, colWidths=[400, 100], hAlign="RIGHT")
            tabela_totais.setStyle(TableStyle([
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("LINEABOVE", (0, -1), (-1, -1), 0.5, colors.black),
                ("TOPPADDING", (0, -1), (-1, -1), 6),
            ]))
            story.append(tabela_totais)
            story.append(Spacer(1, 20))
            story.append(Paragraph("<b>Observações:</b>", styles["Heading5"]))
            story.append(Paragraph(observacoes or "Nenhuma", estilo_normal))
            story.append(Spacer(1, 30))
            story.append(Spacer(1, 40))
            story.append(Paragraph("_____________________________________<br/>Assinatura do Representante", estilo_normal))
            story.append(Spacer(1, 25))
            doc.build(story)
            messagebox.showinfo("Sucesso", f"PDF gerado com sucesso:\n{caminho_pdf}")
            os.startfile(caminho_pdf)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar PDF:\n{e}")

    def importar_dados(self, tipo="clientes"):
    
                caminho = filedialog.askopenfilename(
                    filetypes=[("Planilhas", "*.xlsx *.csv"), ("Todos", "*.*")]
                )
                if not caminho:
                    return

                try:
                    # --- ler arquivo (xlsx ou csv) ---
                    if caminho.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
                        wb = openpyxl.load_workbook(caminho, data_only=True)
                        ws = wb.active
                        rows = list(ws.iter_rows(values_only=True))
                        if not rows:
                            messagebox.showerror("Erro", "Arquivo vazio")
                            return
                        header = [ (str(c) if c is not None else "").strip() for c in rows[0] ]
                        data_rows = [ list(row) for row in rows[1:] ]
                    else:
                        # CSV: tenta descobrir delimitador e lê
                        with open(caminho, newline='', encoding='utf-8') as f:
                            sample = f.read(2048)
                            f.seek(0)
                            try:    
                                dialect = csv.Sniffer().sniff(sample, delimiters=";,")
                                delim = dialect.delimiter
                            except Exception:
                                delim = ';'
                            reader = csv.reader(f, delimiter=delim)
                            header = [ (str(c) if c is not None else "").strip() for c in next(reader) ]
                            data_rows = [row for row in reader]

                    # --- helpers ---
                    def norm(s):
                        return ''.join(c for c in unicodedata.normalize('NFD', s or '') if unicodedata.category(c) != 'Mn').lower().strip()

                    header_norm = [norm(h) for h in header]

                    def find_index(*patterns):
                        for pat in patterns:
                            for i, h in enumerate(header_norm):
                                if pat in h:
                                    return i
                        return None

                    # ---------- importar clientes ----------
                    if tipo == "clientes":
                        idx_razao = find_index('razao', 'cliente', 'empresa', 'nome')
                        idx_cnpj  = find_index('cnpj')
                        idx_tel   = find_index('telefone', 'contato', 'tel')
                        idx_cidade= find_index('cidade')
                        idx_estado= find_index('estado', 'regiao', 'região')

                        if idx_razao is None or idx_cnpj is None:
                            messagebox.showerror("Erro de Coluna", "Arquivo de clientes precisa conter pelo menos as colunas 'Cliente/Razão' e 'CNPJ'.")
                            return

                        importados = 0
                        for row in data_rows:
                            try:
                                razao = str(row[idx_razao]).strip() if idx_razao < len(row) and row[idx_razao] is not None else ''
                                cnpj  = str(row[idx_cnpj]).strip()  if idx_cnpj < len(row) and row[idx_cnpj] is not None else ''
                                telefone = str(row[idx_tel]).strip() if idx_tel is not None and idx_tel < len(row) and row[idx_tel] is not None else ''
                                cidade = str(row[idx_cidade]).strip() if idx_cidade is not None and idx_cidade < len(row) and row[idx_cidade] is not None else ''
                                estado = str(row[idx_estado]).strip() if idx_estado is not None and idx_estado < len(row) and row[idx_estado] is not None else ''

                                if not razao or not cnpj:
                                    continue

                                self.cursor.execute('''
                                    INSERT INTO clientes (razao_social, cnpj, ie, endereco, cidade, estado, cep, telefone, email)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                                ''', (razao, cnpj, '', '', cidade, estado, '', telefone, ''))
                                importados += 1
                            except sqlite3.IntegrityError:
                                # duplicado: ignora
                                continue

                        self.conn.commit()
                        messagebox.showinfo("Sucesso", f"{importados} clientes importados com sucesso!")
                        self.carregar_clientes()
                        self.carregar_combos_pedido()
                        return

                    # ---------- importar produtos ----------
                    if tipo == "produtos":
                        idx_codigo = find_index('cod', 'codigo', 'produto_id', 'id')
                        idx_desc   = find_index('descricao', 'descri', 'produto', 'nome')
                        idx_valor  = find_index('valor', 'preco', 'unitario')
                        idx_tipo   = find_index('tipo')
                        idx_origem = find_index('origem', 'tributacao')
                        idx_ipi    = find_index('ipi')
                        idx_pis    = find_index('pis', 'pis/cofins', 'cofins')
                        idx_icms   = find_index('icms')

                        if idx_codigo is None or idx_desc is None or idx_valor is None:
                            messagebox.showerror("Erro de Coluna", "Arquivo de produtos precisa conter ao menos 'codigo', 'descricao' e 'preco/valor'.")
                            return

                        def to_float_cell(r, idx):
                            if idx is None or idx >= len(r): return 0.0
                            v = r[idx]
                            if v is None: return 0.0
                            s = str(v).replace('R$', '').replace('%', '').replace(',', '.').strip()
                            try:
                                return float(s)
                            except:
                                return 0.0

                        importados = 0
                        for row in data_rows:
                            try:
                                codigo = str(row[idx_codigo]).strip() if idx_codigo is not None and row[idx_codigo] else ''
                                descricao = str(row[idx_desc]).strip() if idx_desc is not None and row[idx_desc] else ''
                                valor_unitario = to_float_cell(row, idx_valor)

                                tipo = str(row[idx_tipo]).strip() if idx_tipo is not None and row[idx_tipo] else ''
                                origem = str(row[idx_origem]).strip() if idx_origem is not None and row[idx_origem] else ''

                                aliq_icms = to_float_cell(row, idx_icms)
                                aliq_ipi  = to_float_cell(row, idx_ipi)
                                aliq_pis  = to_float_cell(row, idx_pis)
                                aliq_cofins = 0.0  # opcional (se quiser separar do PIS/COFINS)

                                if not codigo or not descricao:
                                    continue

                                self.cursor.execute('''
                                    INSERT INTO produtos (codigo, descricao, valor_unitario, tipo, origem_tributacao,
                                                        aliq_icms, aliq_ipi, aliq_pis, aliq_cofins)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                                ''', (codigo, descricao, valor_unitario, tipo, origem,
                                    aliq_icms, aliq_ipi, aliq_pis, aliq_cofins))
                                importados += 1

                            except sqlite3.IntegrityError:
                                # ignora duplicados
                                continue
                            except Exception:
                                # pula linha problemática
                                continue

                        self.conn.commit()
                        messagebox.showinfo("Sucesso", f"{importados} produtos importados com sucesso!")
                        self.carregar_produtos()
                        self.carregar_combos_pedido()
                        return

                    # tipo inválido
                    messagebox.showerror("Tipo inválido", f"Tipo de importação '{tipo}' não suportado.")
                    return

                except Exception as e:  
                    messagebox.showerror("Erro", f"Falha ao importar: {e}")
           
    def exportar_excel_orcamento(self, numero_pedido=None):
        if not numero_pedido:
            item = self.tree_orcamentos.selection()
            if item:
                valores = self.tree_orcamentos.item(item[0], "values")
                numero_pedido = valores[0]
            else:
                messagebox.showerror("Erro", "Nenhum orçamento selecionado.")
                return

        # Buscar dados do orçamento
        self.cursor.execute('''
            SELECT p.data_pedido, c.razao_social, c.cnpj, c.endereco, c.cidade, c.estado,
                p.valor_produtos, p.valor_icms, p.valor_ipi, p.valor_pis, p.valor_cofins,
                p.valor_total, p.representante, p.condicoes_pagamento, p.observacoes, p.validade
            FROM pedidos p
            JOIN clientes c ON p.cliente_id = c.id
            WHERE p.numero_pedido = ?
        ''', (numero_pedido,))
        pedido = self.cursor.fetchone()

        if not pedido:
            messagebox.showerror("Erro", "Orçamento não encontrado.")
            return

        # Preparar dados
        try:
            data_formatada = datetime.strptime(pedido[0], "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%y %H:%M")
        except:
            data_formatada = pedido[0]

        cliente, cnpj, endereco, cidade, estado, subtotal, icms, ipi, pis, cofins, total, representante, cond_pag, obs, validade = (
            pedido[1], pedido[2], pedido[3], pedido[4], pedido[5],
            pedido[6], pedido[7], pedido[8], pedido[9], pedido[10],
            pedido[11], pedido[12], pedido[13], pedido[14], pedido[15]
        )

        endereco_formatado = " - ".join(filter(None, [endereco, cidade, estado])) or "Não informado"

        # Criar planilha
        wb = Workbook()
        ws = wb.active
        ws.title = f"Orçamento {numero_pedido}"

        # Estilos
        header_fill = PatternFill("solid", fgColor="1E3A8A")  # azul escuro
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))

        # Logo (se existir)
        try:
            logo = Image("logo.png")
            logo.width, logo.height = 150, 40
            ws.add_image(logo, "A1")
        except:
            pass

        # Cabeçalho
        ws.merge_cells("A2:E2")
        ws["A2"] = f"ORÇAMENTO Nº {numero_pedido}"
        ws["A2"].font = Font(size=14, bold=True)
        ws["A2"].alignment = Alignment(horizontal="center")

        ws["A4"], ws["B4"] = "Data:", data_formatada
        ws["A5"], ws["B5"] = "Cliente:", cliente
        ws["A6"], ws["B6"] = "CNPJ:", cnpj
        ws["A7"], ws["B7"] = "Endereço:", endereco_formatado
        ws["A8"], ws["B8"] = "Representante:", representante
        ws["A9"], ws["B9"] = "Condições de Pagamento:", cond_pag
        ws["A10"], ws["B10"] = "Validade:", f"{validade} dias"

        # Itens do orçamento
        ws.append([])
        linha_inicio = 12
        colunas = ["Código", "Descrição", "Qtd", "Valor Unitário", "Total Item"]

        for col, nome in enumerate(colunas, start=1):
            cel = ws.cell(row=linha_inicio, column=col, value=nome)
            cel.font = header_font
            cel.fill = header_fill
            cel.alignment = Alignment(horizontal="center")
            cel.border = border

        # Buscar itens
        self.cursor.execute('''
            SELECT pr.codigo, pr.descricao, pi.qtd, pi.valor_unitario
            FROM pedido_itens pi
            JOIN produtos pr ON pi.produto_id = pr.id
            WHERE pi.numero_pedido = ?
        ''', (numero_pedido,))
        itens = self.cursor.fetchall()

        linha = linha_inicio + 1
        subtotal_calc = 0

        for cod, desc, qtd, v_unit in itens:
            v_total_item = qtd * v_unit
            subtotal_calc += v_total_item
            dados = [cod, desc, qtd, v_unit, v_total_item]

            for col, valor in enumerate(dados, start=1):
                cel = ws.cell(row=linha, column=col, value=valor)
                if col >= 3:  # qtd, unitário e total
                    cel.number_format = 'R$ #,##0.00'
                    cel.alignment = Alignment(horizontal="right")
                cel.border = border

            # Linhas alternadas (efeito zebra)
            if (linha % 2) == 0:
                for col in range(1, 6):
                    ws.cell(row=linha, column=col).fill = PatternFill("solid", fgColor="F9F9F9")

            linha += 1

        # Resumo de totais
        linha += 1
        totais = [
            ("Subtotal:", subtotal_calc),
            ("ICMS:", icms),
            ("IPI:", ipi),
            ("PIS:", pis),
            ("COFINS:", cofins)
        ]

        desconto = float(self.entry_desconto.get() or 0)
        if desconto > 0:
            totais.append(("Desconto:", -desconto))

        total_geral = subtotal_calc + icms + ipi + pis + cofins - desconto
        totais.append(("TOTAL GERAL", total_geral))

        for label, valor in totais:
            ws.merge_cells(start_row=linha, start_column=4, end_row=linha, end_column=4)
            cel_label = ws.cell(row=linha, column=4, value=label)
            cel_label.alignment = Alignment(horizontal="right")
            cel_label.font = Font(bold=True)
            cel_val = ws.cell(row=linha, column=5, value=valor)
            cel_val.number_format = 'R$ #,##0.00'
            if "TOTAL" in label:
                cel_label.fill = PatternFill("solid", fgColor="124A12")
                cel_label.font = Font(bold=True, size=12, color="FFFFFF")
                cel_val.font = Font(bold=True, size=12, color="FFFFFF")
                cel_val.fill = PatternFill("solid", fgColor="124A12")
            linha += 1

        # Observações
        if obs:
            linha += 1
            ws.merge_cells(f"A{linha}:E{linha+2}")
            cel = ws[f"A{linha}"]
            cel.value = f"Observações: {obs}"
            cel.alignment = Alignment(wrap_text=True, vertical="top")
            cel.fill = PatternFill("solid", fgColor="EFEFEF")

        # Espaço para assinatura
        linha += 4
        ws.merge_cells(f"A{linha}:C{linha}")
        ws[f"A{linha}"] = "Assinatura do Cliente: __________________________"

        # Rodapé
        linha += 2
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=5)
        ws.cell(row=linha, column=1,
                value=f"Gerado em {datetime.now().strftime('%d/%m/%y %H:%M')} - Sistema Interno de Orçamentos"
            ).alignment = Alignment(horizontal="center")

        # Autoajuste das colunas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Salvar
        nome_arquivo = f"Orcamento_{numero_pedido}.xlsx"
        wb.save(nome_arquivo)
        os.startfile(nome_arquivo)

    def abrir_formulario_empresa(self, empresa=None):
        top = tk.Toplevel(self.root)
        top.title("Cadastro de Empresa")
        top.geometry("650x420")
        # Estrutura de campos
        labels = ['Nome:', 'CNPJ:', 'IE:', 'CEP:', 'Endereço:', 'Cidade:', 'Estado:', 'Telefone:', 'Email:']
        entries = {}

        for i, label in enumerate(labels):
            ttk.Label(top, text=label).grid(row=i//2, column=(i%2)*2, sticky='w', padx=5, pady=5)
            entry = ttk.Entry(top, width=28)
            entry.grid(row=i//2, column=(i%2)*2+1, padx=5, pady=5, sticky="ew")
            chave = normalizar_chave(label)
            entries[chave] = entry

        # 🔍 Busca automática do endereço pelo CEP
        entry_cep = entries["cep"]
        entry_cep.bind("<FocusOut>", lambda e: self.buscar_cep(entry_cep.get(), entries))

        # Preenche campos se for edição
        if empresa:
            keys = ['id', 'nome', 'cnpj', 'ie', 'cep', 'endereco', 'cidade', 'estado', 'telefone', 'email']
            for k, v in zip(keys, empresa):
                if k in entries:
                    entries[k].insert(0, v or "")

        def salvar():
            dados = {k: v.get().strip() for k, v in entries.items()}
            campos_obrigatorios = {"nome": "Nome", "cnpj": "CNPJ"}
            faltando = [nome for chave, nome in campos_obrigatorios.items() if not dados.get(chave)]

            if faltando:
                messagebox.showwarning("Atenção", f"Preencha os campos obrigatórios: {', '.join(faltando)}")
                return

            try:
                if empresa:
                    self.cursor.execute('''
                        UPDATE empresas
                        SET nome=?, cnpj=?, ie=?, cep=?, endereco=?, cidade=?, estado=?, telefone=?, email=?
                        WHERE id=?
                    ''', (
                        dados['nome'], dados['cnpj'], dados.get('ie'), dados.get('cep'), dados.get('endereco'),
                        dados.get('cidade'), dados.get('estado'), dados.get('telefone'), dados.get('email'), empresa[0]
                    ))
                else:
                    self.cursor.execute('''
                        INSERT INTO empresas (nome, cnpj, ie, cep, endereco, cidade, estado, telefone, email)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        dados['nome'], dados['cnpj'], dados.get('ie'), dados.get('cep'), dados.get('endereco'),
                        dados.get('cidade'), dados.get('estado'), dados.get('telefone'), dados.get('email')
                    ))
                self.conn.commit()
                self.carregar_empresas()
                top.destroy()
            except sqlite3.IntegrityError:
                messagebox.showerror("Erro", "CNPJ já cadastrado!")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar empresa: {e}")

        ttk.Button(top, text="Salvar", bootstyle=SUCCESS, command=salvar).grid(row=6, column=0, columnspan=2, pady=10)

    def carregar_empresas(self):
        try:
            for item in getattr(self, 'tree_empresas', []).get_children():
                self.tree_empresas.delete(item)
        except Exception:
            pass

        # Continua buscando o ID (para identificar ao editar/excluir)
        self.cursor.execute("SELECT id, nome, cnpj, cidade, telefone FROM empresas")
        for row in self.cursor.fetchall():
            id_, nome, cnpj, cidade, telefone = row
            cnpj_fmt = formatar_cnpj(cnpj)
            telefone_fmt = formatar_telefone(telefone)
            # Guardamos o ID como "tags" (oculto), mas não mostramos na tabela
            self.tree_empresas.insert('', 'end', values=(nome, cnpj_fmt, cidade, telefone_fmt), tags=(id_,))

    def editar_empresa(self):
        sel = self.tree_empresas.selection()
        if not sel:
            messagebox.showwarning("Atenção", "Selecione uma empresa.")
            return
        vals = self.tree_empresas.item(sel[0], 'values')
        empresa_id = self.tree_empresas.item(sel[0], 'tags')[0]
        self.cursor.execute("SELECT * FROM empresas WHERE id=?", (empresa_id,))
        empresa = self.cursor.fetchone()
        if empresa:
            self.abrir_formulario_empresa(empresa)

    def excluir_empresa(self):
        sel = self.tree_empresas.selection()
        if not sel:
            messagebox.showwarning("Atenção", "Selecione uma empresa para excluir.")
            return
        if not messagebox.askyesno("Confirmar", "Deseja excluir a(s) empresa(s) selecionada(s)?"):
            return
        try:
            for item in sel:
                vals = self.tree_empresas.item(item, 'values')
                self.cursor.execute("DELETE FROM empresas WHERE id=?", (vals[0],))
            self.conn.commit()
            self.carregar_empresas()
            self.carregar_combos_pedido()
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao excluir: {e}")
    def __del__(self):
            if hasattr(self, 'conn'):
                self.conn.close()

if __name__ == "__main__":      
    root = tb.Window(themename="solar")
    app = SistemaPedidos(root)
    root.mainloop()
    